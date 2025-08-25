import openpyxl as px

def excel_observation_cheker(excel_path: str, texts: list[str]) -> set[str]:
    """
    Inserta en bloque los textos en la columna G de 'Matriz Obs' desde la fila 13.
    Devuelve un set con los textos NORMALIZADOS que fueron agregados (no existentes previamente).
    """
    SHEET_NAME = "Matriz Obs"
    COL = 7
    START_ROW = 13

    def _norm(s: str) -> str:
        return (s or "").strip()

    def _first_empty_row(ws, col=COL, start=START_ROW):
        mr = ws.max_row
        if mr < start:
            return start
        r = mr
        while r >= start:
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip() != "":
                return r + 1
            r -= 1
        return start

    if not texts:
        return set()

    #Abre el excel
    wb = px.load_workbook(filename=excel_path, data_only=True, read_only=False)
    ws = wb[SHEET_NAME]

    # Observaciones ya existentes
    existentes = {
        _norm(v)
        for (v,) in ws.iter_rows(
            min_row=START_ROW,
            max_row=ws.max_row,
            min_col=COL,
            max_col=COL,
            values_only=True
        )
        if v not in (None, "")
    }

    # Filtra textos nuevos
    vistos = set()
    to_add = []
    for t in texts:
        tn = _norm(t)
        if not tn or tn in existentes or tn in vistos:
            continue
        to_add.append(t)
        vistos.add(tn)

    agregados = set()
    if to_add:
        row = _first_empty_row(ws, COL, START_ROW)
        for t in to_add:
            ws.cell(row=row, column=COL, value=t)
            agregados.add(_norm(t))
            row += 1
        wb.save(excel_path)

    wb.close()
    return agregados



"""
EXCEL_PATHS = ["docs/BIM_20250508_EDMAX_BIM_ANI_Matriz_de_observaciones.xlsx", 
               "docs/ELECTRICA_20250508_EDMAX_ELE_ANI_Matriz_de_observaciones.xlsx",
               "docs/GEOL Y ESTRUCT_20250508_EDMAX_EST Y GEOL_ANI_Matriz_de_observaciones.xlsx",
               "docs/GEOMETRICO_20250327_EDMAX_ANI_Matriz_de_observaciones_Geometrico.xlsx",
               "docs/PAVIMENTOS_20250705_EDMAX_ANI_Matriz_de_observaciones.xlsx",
               "docs/SEG VIAL_20250327_EDMAX_ANI_Matriz_de_observaciones_SGV.xlsx",
               "docs/TOPOGRAFIA _20250327_EDMAX_ANI_Matriz_de_observaciones.xlsx",
               "docs/TRANSITO_20250507_EDMAX_TRAN_ANI_Matriz_de_observaciones.xlsx"
               ]

"""


"""
import json
#TODO: borras esto
# Ruta al archivo JSON
ruta_json = "docs/output.json"
# Abrir y leer el JSON
with open(ruta_json, "r", encoding="utf-8") as f:
    datos = json.load(f)


#Saca los valores unicos de especialidad siempre y cuando sea observacion
especialidades = set()
for p in datos["resultados"]:   
    if p.get("especialidad") and p.get("especialidad") != "DESCONOCIDA" and p.get("etiqueta", "").lower() == "observacion":
        especialidades.add(p.get("especialidad"))   

# Itera por las especialidades y agrega las observaciones
for esp in especialidades:
    #Obtiene la URL del excel correspondiente a la especialidad
    if esp == "BIM":
        excel_path = EXCEL_PATHS[0]
    elif esp == "ELECTRICA":  
        excel_path = EXCEL_PATHS[1]
    elif esp == "GEOL Y ESTRUCT":
        excel_path = EXCEL_PATHS[2] 
    elif esp == "GEOMETRICO":
        excel_path = EXCEL_PATHS[3]
    elif esp == "PAVIMENTOS":     
        excel_path = EXCEL_PATHS[4]
    elif esp == "SEG VIAL":
        excel_path = EXCEL_PATHS[5]
    elif esp == "TOPOGRAFIA":
        excel_path = EXCEL_PATHS[6]
    elif esp == "TRANSITO":
        excel_path = EXCEL_PATHS[7]


    #Obtiene todos los textos que pertenecen a la especialidad y que sean observaciones
    textos = [p["texto"] for p in datos["resultados"] if p.get("especialidad") == esp and p.get("etiqueta", "").lower() == "observacion"]

    #Abre la matriz de observaciones de la especialidad y agrega las observaciones
    obs_agregadas = excel_observation_cheker(excel_path, textos)  


    for p in datos["resultados"]:
        if p.get("especialidad") == esp and p.get("etiqueta", "").lower() == "observacion":
            p["observacion_agregada"] = _norm(p["texto"]) in obs_agregadas 
    

a = 10
"""
